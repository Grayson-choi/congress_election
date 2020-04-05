from selenium import webdriver
import time
import urllib.request as req
from bs4 import BeautifulSoup
import urllib.parse as par
import os
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
import urllib.request as req
import os
import openpyxl
from datetime import datetime
import re

def sgg_crawl():
    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    options.add_argument('window-size=1920x1080')
    options.add_argument("disable-gpu")

    browser = webdriver.Chrome("/Users/jw/Desktop/study/project/congress_election/chromedriver", options=options)
    # browser = webdriver.Chrome("./chromedriver")
    # 윈도우는 아래 코드를 실행
    # browser = webdriver.Chrome("./chromedriver.exe")

    # 선거구 및 읍면동현황 이
    browser.get("http://info.nec.go.kr/main/showDocument.xhtml?electionId=0020200415&topMenuId=BI&secondMenuId=BIGI05")
    time.sleep(4)

    #국회의원 선거 클릭
    browser.find_element_by_css_selector("#electionId2").click()
    time.sleep(2)

    #도시 선택 cityCode
    city_list = browser.find_elements_by_css_selector("select#cityCode > option")
    city_list = list(map(lambda x: x.text, city_list))
    city_list = ['서울특별시', '부산광역시', '대구광역시', '인천광역시', '광주광역시', '대전광역시', '울산광역시', '세종특별자치시', '경기도', '강원도', '충청북도', '충청남도', '전라북도', '전라남도', '경상북도', '경상남도', '제주특별자치도']
    print(len(city_list))
    dong_list = []

    confirm_sgg = ""

    for city in city_list:
        # 도시 선택
        select_city = Select(browser.find_element_by_id('cityCode'))

        #서울 특별시 선택
        select_city.select_by_visible_text(city)
        time.sleep(1)

        # 조회 버튼 선택 (홈페이지에서 img 타입의 버튼이 하나라서 이렇게 씀)
        browser.find_element_by_css_selector("#spanSubmit").click()

        # 긁어오는 코드
        cnt = 1
        while True:
            try:
                sgg = browser.find_element_by_css_selector(f"#table01 > tbody > tr:nth-child({cnt}) > td.alignL.rowspan").text
                sigun = browser.find_element_by_css_selector(f"#table01 > tbody > tr:nth-child({cnt}) > td:nth-child(3)").text
                dong = browser.find_element_by_css_selector(f"#table01 > tbody > tr:nth-child({cnt}) > td:nth-child(4)").text
                # print("sgg: " + sgg.text)
                # print("sigun: " + sigun.text)
                # print(dong.text)
                print("")
                for i in dong.split(", "):
                    dic = {}
                    if not sgg:
                        sgg = confirm_sgg
                    dic["city"] = city
                    dic["sigun"] = sigun
                    dic["dong"] = i
                    dic["sgg"] = sgg
                    dong_list.append(dic)
                    confirm_sgg = sgg
                cnt += 1
            except:
                break
        print(city+"============================")
        print(dong_list)

        browser.get("http://info.nec.go.kr/main/showDocument.xhtml?electionId=0020200415&topMenuId=BI&secondMenuId=BIGI05")
        time.sleep(4)

        # 국회의원 선거 클릭
        browser.find_element_by_css_selector("#electionId2").click()
        time.sleep(2)

    return dong_list

#
# dong_list = sgg_crawl()
#
# # 엑셀 파일 있는지 먼저 확인
# excel_dir = "./멜론_크롤링.xlsx"
#
# if not os.path.exists(excel_dir):  # 파일이 존재하지 않으면, 파일 생성
#     openpyxl.Workbook().save(excel_dir)
#
# # 엑셀 파일 불러오기
# book = openpyxl.load_workbook(excel_dir)
# # 새로운 시트 생성, 시트 이름은 현재 날짜,시간(**년**월**일**시**분**초)
# sheet = book.create_sheet()
# now = datetime.now()
# sheet.title = "{}년{}월{}일{}시{}분{}초".format(now.year, now.month, now.day, now.hour, now.minute, now.second)
# # 불필요한 시트는 지운다.
# if "Sheet" in book.sheetnames:
#     book.remove(book["Sheet"])
# # 열너비 조절
# sheet.column_dimensions["A"].width = 15
# sheet.column_dimensions["B"].width = 30
# sheet.column_dimensions["C"].width = 20
#
# for i in range(len(title)):
#     # 행높이 조절
#     sheet.row_dimensions[i + 1].height = 92.4
#
#     img_file_name = img_dir + "/" + re.sub("[\\\/:*?\"<>\|]", " ", title[i].string) + ".png"
#
#     req.urlretrieve(img[i].attrs["src"], img_file_name)  # 앨범 이미지 저장
#     print("{}위 : {} - {}".format(i + 1, title[i].string, name[i].text))
#
#     # 앨범이미지 추가
#     img_for_excel = Image(img_file_name)
#     sheet.add_image(img_for_excel, "A" + str(i + 1))
#     # 노래제목 추가
#     sheet.cell(row=i + 1, column=2).value = title[i].string
#     # 가수이름 추가
#     sheet.cell(row=i + 1, column=3).value = name[i].text
#     # 앨범이름 추가
#     sheet.cell(row=i + 1, column=4).value = album[i].string
#     # 한 행 기록할 때마다 바로 저장
#     book.save(excel_dir)
#

    # cnt = 0
    # while True:
    #     try:
    #         cnt += 1
    #         sgg = browser.find_element_by_css_selector(f"#table01 > tbody > tr:nth-child({cnt}) > td.alignL.rowspan")
    #         dong = browser.find_element_by_css_selector(f"#table01 > tbody > tr:nth-child({cnt}) > td:nth-child(4)")
    #         dong_list.append(sgg.text)
    #         dong_list.append(dong.text)
    #     except:
    #         break

    # print(dong_list)

    # sgg_list2 = list(map(lambda x: x.text, sgg_list2))
    # print(sgg_list2)
    # for i in sgg_list2:
    #     print(i.text)
    #table01 > tbody > tr:nth-child(1) > td:nth-child(4)

if __name__ == "__main__":
    dong_list = sgg_crawl()
